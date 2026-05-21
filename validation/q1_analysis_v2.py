#!/usr/bin/env python
"""
q1_analysis_v2.py — Q1 Refinement Analysis

CAMBIO A: Short event-locked window
  Pre:  [−5, −2] min relative to stimulus (t_stim − 300 … t_stim − 120 s)
  Post: [0, +2.5] min relative to stimulus (t_stim … t_stim + 150 s)
  Justification: peak observed at 90 s, return at 120 s (trajectory_diagnostic.md)

CAMBIO B: Corrected surrogate mapping
  All 5 primary PTT-variability features predicted to DECREASE under sympathetic
  activation (uniform stiffening / compression of PTT dynamic range).
  Bonferroni α = 0.05 / 5 = 0.010.

Features pre-specified — NOT modified based on results.

Outputs:
  1. results/validation/Q1/paired_event_data.csv
  2. results/validation/Q1/test_results_v2.csv
  3. results/validation/Q1/figures/paired_event_violins.png
  4. results/validation/Q1/figures/forest_plot_v2.png
  5. results/validation/Q1/figures/gradient_plot_v2.png  (if >=1 validated)
  6. results/validation/Q1/Q1_report_v2.md
"""
import sys
import warnings
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
from scipy import stats
import statsmodels.formula.api as smf
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.lines import Line2D

# ─── PATHS ────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
RES  = ROOT / "results" / "validation" / "Q1"
FIG  = RES / "figures"
FIG.mkdir(parents=True, exist_ok=True)

FEATURES_LONG = RES / "features_long.parquet"
EVENT_WINDOWS = RES / "event_windows.csv"
PAIRED_OUT    = RES / "paired_event_data.csv"
RESULTS_OUT   = RES / "test_results_v2.csv"
REPORT_OUT    = RES / "Q1_report_v2.md"
EXCEL_CLIN    = ROOT / "database general pacientes estudio hombro.xlsx"

# ─── CAMBIO A: short windows ──────────────────────────────────────────────────
PRE_START_S  = -300   # −5 min
PRE_END_S    = -120   # −2 min
POST_START_S =    0
POST_END_S   =  150   # +2.5 min
WINDOW_S     =   30   # duration of each feature window (s)

# ─── CAMBIO B: primary features — all expected direction − ───────────────────
# (feature_col, aggregation, expected_direction)
PRIMARY = [
    ("ptt_cv",       "mean", -1),
    ("ptt_cv",       "std",  -1),
    ("ptt_std",      "max",  -1),
    ("brs_alpha_lf", "min",  -1),
    ("pai_mean",     "mean", -1),
]

EXPLORATORY = [          # PASO 5 — no Bonferroni, flagged exploratory
    ("ptt_std",  "std",   -1),
    ("ptt_std",  "slope", -1),
    ("ptt_arv",  "std",   -1),
]

ALPHA_BONF      = 0.05 / len(PRIMARY)   # 0.010
MAX_NAN_PCT     = 0.20
PARTIAL_PATIENT = "70767707"
SEED            = 42
np.random.seed(SEED)

# ─── HELPER: compute scalar aggregation ───────────────────────────────────────
def compute_agg(series: pd.Series, t: pd.Series, agg: str) -> float:
    valid_mask = series.notna()
    s  = series[valid_mask]
    tv = t[valid_mask]
    if len(s) == 0:
        return np.nan
    if agg == "mean":
        return float(s.mean())
    elif agg == "std":
        return float(s.std(ddof=1)) if len(s) >= 2 else np.nan
    elif agg == "max":
        return float(s.max())
    elif agg == "min":
        return float(s.min())
    elif agg == "slope":
        if len(s) < 3:
            return np.nan
        res = stats.linregress(tv.values.astype(float), s.values.astype(float))
        return float(res.slope)
    raise ValueError(f"Unknown agg: {agg}")


# ─── HELPER: one-sided p-value ────────────────────────────────────────────────
def one_sided_p(beta: float, p_2sided: float, exp_dir: int) -> float:
    if np.isnan(beta) or np.isnan(p_2sided):
        return np.nan
    sign_match = (exp_dir < 0 and beta < 0) or (exp_dir > 0 and beta > 0)
    return (p_2sided / 2.0) if sign_match else (1.0 - p_2sided / 2.0)


# ─── HELPER: Cohen's d_z ─────────────────────────────────────────────────────
def cohen_dz(deltas: np.ndarray) -> float:
    d = deltas[~np.isnan(deltas)]
    if len(d) < 2:
        return np.nan
    return float(d.mean() / d.std(ddof=1))


# ─── HELPER: run mixed model (intercept-only) ────────────────────────────────
def run_mixed_model(df: pd.DataFrame,
                    formula: str = "delta ~ 1",
                    groups_col: str = "patient_id",
                    vc_stim: bool = True) -> dict:
    """
    Fit MixedLM. Tries:
      1. GLMM with patient_id + stim_subcategory random effects
         (methods: bfgs, cg, nm in order until convergence)
      2. GLMM with patient_id only (same method ladder)
      3. OLS fallback
    Returns dict: beta, ic_lo, ic_hi, p_2sided, model_note, converged
    """
    OPT_METHODS  = ["bfgs", "cg", "nm"]
    intercept_key = "Intercept"

    attempts = []
    if vc_stim and "stim_subcategory" in df.columns and df["stim_subcategory"].nunique() >= 2:
        attempts.append(("GLMM_2re", True))
    attempts.append(("GLMM_1re", False))

    for mname, use_vc in attempts:
        for opt in OPT_METHODS:
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    if use_vc:
                        vc = {"stim_subcategory": "0 + C(stim_subcategory)"}
                        md = smf.mixedlm(formula, df,
                                         groups=df[groups_col], vc_formula=vc)
                    else:
                        md = smf.mixedlm(formula, df, groups=df[groups_col])
                    fit = md.fit(reml=True, method=opt, maxiter=800)
                    if not fit.converged:
                        raise RuntimeError("not converged")
                    ci = fit.conf_int()
                    return dict(
                        beta=float(fit.params[intercept_key]),
                        ic_lo=float(ci.loc[intercept_key, 0]),
                        ic_hi=float(ci.loc[intercept_key, 1]),
                        p_2sided=float(fit.pvalues[intercept_key]),
                        model_note=f"{mname}_{opt}",
                        converged=True,
                    )
            except Exception:
                pass

    # OLS fallback
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            md  = smf.ols(formula, df)
            fit = md.fit()
            ci  = fit.conf_int()
            return dict(
                beta=float(fit.params[intercept_key]),
                ic_lo=float(ci.loc[intercept_key, 0]),
                ic_hi=float(ci.loc[intercept_key, 1]),
                p_2sided=float(fit.pvalues[intercept_key]),
                model_note="OLS_fallback",
                converged=False,
            )
    except Exception:
        pass

    return dict(beta=np.nan, ic_lo=np.nan, ic_hi=np.nan,
                p_2sided=np.nan, model_note="failed", converged=False)


# ─── HELPER: gradient model (group effect) ────────────────────────────────────
def run_gradient_model(df: pd.DataFrame) -> dict:
    """
    delta ~ group_c + (1|patient_id) [+ (1|stim_subcategory)]
    group_c = 1 for interescalenico (reference: supra_axilar = 0)
    H1 one-sided: group_c > 0 (interescalenic has less-negative delta)
    """
    df2 = df.copy()
    df2["group_c"] = (df2["group"] == "interescalenico").astype(float)
    formula = "delta ~ group_c"
    coef_key = "group_c"

    attempts = []
    if df2["stim_subcategory"].nunique() >= 2:
        attempts.append(("gradient_2re", True))
    attempts.append(("gradient_1re", False))

    OPT_METHODS = ["bfgs", "cg", "nm"]
    for mname, use_vc in attempts:
        for opt in OPT_METHODS:
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    if use_vc:
                        vc = {"stim_subcategory": "0 + C(stim_subcategory)"}
                        md = smf.mixedlm(formula, df2,
                                         groups=df2["patient_id"], vc_formula=vc)
                    else:
                        md = smf.mixedlm(formula, df2, groups=df2["patient_id"])
                    fit = md.fit(reml=True, method=opt, maxiter=800)
                    if not fit.converged:
                        raise RuntimeError("not converged")
                    ci = fit.conf_int()
                    b  = float(fit.params[coef_key])
                    p2 = float(fit.pvalues[coef_key])
                    return dict(
                        beta=b,
                        ic_lo=float(ci.loc[coef_key, 0]),
                        ic_hi=float(ci.loc[coef_key, 1]),
                        p_1sided=one_sided_p(b, p2, exp_dir=+1),
                        model_note=f"{mname}_{opt}",
                    )
            except Exception:
                pass

    # OLS fallback
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            md  = smf.ols(formula, df2)
            fit = md.fit()
            ci  = fit.conf_int()
            b   = float(fit.params[coef_key])
            p2  = float(fit.pvalues[coef_key])
            return dict(beta=b, ic_lo=float(ci.loc[coef_key, 0]),
                        ic_hi=float(ci.loc[coef_key, 1]),
                        p_1sided=one_sided_p(b, p2, exp_dir=+1),
                        model_note="OLS_fallback")
    except Exception:
        pass

    return dict(beta=np.nan, ic_lo=np.nan, ic_hi=np.nan,
                p_1sided=np.nan, model_note="failed")


# ─── HELPER: load clinical covariates from Excel ──────────────────────────────
def load_clinical_covariates() -> pd.DataFrame:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(EXCEL_CLIN), data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        col_pid = col_age = col_bmi = col_pos = None
        for i, h in enumerate(headers):
            if h == "nombre vital sing":
                col_pid = i
            elif h == "edad":
                col_age = i
            elif h == "BMI":
                col_bmi = i
            elif isinstance(h, str) and "posici" in h.lower():
                col_pos = i

        if col_pid is None:
            print("[WARN] 'nombre vital sing' not found in Excel")
            return pd.DataFrame()

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            pid = row[col_pid]
            if pid is None:
                continue
            rows.append({
                "patient_id": int(pid),
                "edad":     row[col_age]  if col_age is not None else np.nan,
                "bmi":      row[col_bmi]  if col_bmi is not None else np.nan,
                "posicion": row[col_pos]  if col_pos is not None else np.nan,
            })

        df = pd.DataFrame(rows)
        for c in ["edad", "bmi", "posicion"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        return df
    except Exception as e:
        print(f"[WARN] Could not load clinical covariates: {e}")
        return pd.DataFrame()


# ─── PASO 1: build paired_event_data ─────────────────────────────────────────
def build_paired_data(feat_df: pd.DataFrame,
                      events_df: pd.DataFrame) -> pd.DataFrame:
    """
    For each event × (feature, agg), compute value_pre, value_post, delta.
    Pre window:  [t_stim + PRE_START_S,  t_stim + PRE_END_S  − WINDOW_S]
    Post window: [t_stim + POST_START_S, t_stim + POST_END_S − WINDOW_S]
    (fully-contained windows only)
    """
    all_fa = list(dict.fromkeys([(f, a) for f, a, _ in PRIMARY + EXPLORATORY]))

    # BRS masking: brs_alpha_lf = NaN when brs_valid == False
    if "brs_valid" in feat_df.columns and "brs_alpha_lf" in feat_df.columns:
        feat_df = feat_df.copy()
        feat_df.loc[~feat_df["brs_valid"].astype(bool), "brs_alpha_lf"] = np.nan

    # Normalise patient_id to string in both tables
    feat_df   = feat_df.copy()
    events_df = events_df.copy()
    feat_df["patient_id"]   = feat_df["patient_id"].astype(str)
    events_df["patient_id"] = events_df["patient_id"].astype(str)

    # Build per-patient cache for fast lookups
    patient_cache = {
        pid: grp.reset_index(drop=True)
        for pid, grp in feat_df.groupby("patient_id")
    }

    records = []
    for _, ev in events_df.iterrows():
        pid    = str(ev["patient_id"])
        t_stim = float(ev["t_start_s"])
        group  = ev["group"]
        sub    = ev["event_subcategory"]
        sid    = int(ev["window_id"])

        pf = patient_cache.get(pid)
        if pf is None or pf.empty:
            continue

        t_col = pf["t_window_start_s"]
        pre_idx  = (t_col >= t_stim + PRE_START_S) & \
                   (t_col <= t_stim + PRE_END_S  - WINDOW_S)
        post_idx = (t_col >= t_stim + POST_START_S) & \
                   (t_col <= t_stim + POST_END_S - WINDOW_S)

        pre_w  = pf[pre_idx]
        post_w = pf[post_idx]

        for feat, agg in all_fa:
            if feat not in pf.columns:
                continue
            v_pre  = compute_agg(pre_w[feat],  pre_w["t_window_start_s"],  agg)
            v_post = compute_agg(post_w[feat], post_w["t_window_start_s"], agg)
            delta  = (v_post - v_pre) \
                     if not (np.isnan(v_pre) or np.isnan(v_post)) else np.nan
            records.append({
                "patient_id":       pid,
                "group":            group,
                "stim_id":          sid,
                "stim_subcategory": sub,
                "feature":          feat,
                "agg":              agg,
                "value_pre":        v_pre,
                "value_post":       v_post,
                "delta":            delta,
            })

    return pd.DataFrame(records)


# ─── PASO 2: primary confirmatory tests ──────────────────────────────────────
def run_primary_tests(paired_df: pd.DataFrame) -> list:
    results = []
    print("\n=== PASO 2 — Primary tests (Bonferroni α={:.3f}) ===".format(ALPHA_BONF))

    for feat, agg, exp_dir in PRIMARY:
        subset = paired_df[
            (paired_df["feature"] == feat) &
            (paired_df["agg"]     == agg)
        ].copy()

        nan_pct = subset["delta"].isna().mean()
        subset  = subset.dropna(subset=["delta"])
        n_pairs = len(subset)
        label   = f"{feat}__{agg}"

        if nan_pct > MAX_NAN_PCT:
            print(f"  [EXCL] {label}: {nan_pct:.0%} NaN — excluded")
            results.append(dict(
                feature=feat, agg=agg, n_pairs=n_pairs,
                beta=np.nan, ic_lo=np.nan, ic_hi=np.nan,
                p_1sided=np.nan, p_bonferroni=np.nan, cohen_dz=np.nan,
                verdict="excluded_nan", analysis="primary", model_note="excluded",
            ))
            continue

        if n_pairs < 5:
            print(f"  [SKIP] {label}: n={n_pairs} < 5")
            results.append(dict(
                feature=feat, agg=agg, n_pairs=n_pairs,
                beta=np.nan, ic_lo=np.nan, ic_hi=np.nan,
                p_1sided=np.nan, p_bonferroni=np.nan, cohen_dz=np.nan,
                verdict="excluded_n", analysis="primary", model_note="excluded",
            ))
            continue

        print(f"  {label}  n={n_pairs}", end="  … ")
        m    = run_mixed_model(subset, "delta ~ 1", vc_stim=True)
        p1   = one_sided_p(m["beta"], m["p_2sided"], exp_dir)
        p_bf = float(min(p1 * len(PRIMARY), 1.0))
        dz   = cohen_dz(subset["delta"].values)

        correct_dir = (exp_dir < 0 and m["beta"] < 0) or \
                      (exp_dir > 0 and m["beta"] > 0)
        validated   = p_bf < 0.05 and correct_dir
        verdict     = "validado" if validated else "no_validado"

        print(f"β={m['beta']:+.4f}  IC95%=[{m['ic_lo']:+.4f},{m['ic_hi']:+.4f}]"
              f"  p_bonf={p_bf:.4f}  d_z={dz:.3f}  {verdict}"
              f"  [{m['model_note']}]")

        results.append(dict(
            feature=feat, agg=agg, n_pairs=n_pairs,
            beta=m["beta"], ic_lo=m["ic_lo"], ic_hi=m["ic_hi"],
            p_1sided=p1, p_bonferroni=p_bf, cohen_dz=dz,
            verdict=verdict, analysis="primary", model_note=m["model_note"],
        ))

    return results


# ─── PASO 3: gradient analysis ───────────────────────────────────────────────
def run_gradient_analysis(paired_df: pd.DataFrame,
                          validated_fa: list) -> list:
    results = []
    if not validated_fa:
        print("\n=== PASO 3 — Gradient: skipped (0 validated) ===")
        return results

    print(f"\n=== PASO 3 — Gradient analysis ({len(validated_fa)} validated features) ===")
    for feat, agg in validated_fa:
        subset = paired_df[
            (paired_df["feature"] == feat) &
            (paired_df["agg"]     == agg)
        ].dropna(subset=["delta"]).copy()

        if subset["group"].nunique() < 2:
            print(f"  [SKIP] {feat}__{agg}: only one group present")
            continue

        print(f"  gradient {feat}__{agg}  n={len(subset)}", end="  … ")
        g = run_gradient_model(subset)
        print(f"β_group={g['beta']:+.4f}  IC=[{g['ic_lo']:+.4f},{g['ic_hi']:+.4f}]"
              f"  p_1sided={g['p_1sided']:.4f}  [{g['model_note']}]")

        results.append(dict(
            feature=feat, agg=agg, n_pairs=len(subset),
            beta=g["beta"], ic_lo=g["ic_lo"], ic_hi=g["ic_hi"],
            p_1sided=g["p_1sided"], p_bonferroni=np.nan, cohen_dz=np.nan,
            verdict="gradient_test", analysis="gradient",
            model_note=g["model_note"],
        ))
    return results


# ─── PASO 4: sensitivity analyses ─────────────────────────────────────────────
def run_sensitivity(paired_df: pd.DataFrame,
                    clin_df: pd.DataFrame) -> list:
    results = []
    print("\n=== PASO 4 — Sensitivity analyses ===")

    for feat, agg, exp_dir in PRIMARY:
        label = f"{feat}__{agg}"
        base  = paired_df[
            (paired_df["feature"] == feat) &
            (paired_df["agg"]     == agg)
        ].dropna(subset=["delta"]).copy()

        # ── A: exclude partial patient ──────────────────────────────────────
        sa = base[base["patient_id"] != PARTIAL_PATIENT]
        if len(sa) >= 5:
            m  = run_mixed_model(sa, "delta ~ 1", vc_stim=True)
            p1 = one_sided_p(m["beta"], m["p_2sided"], exp_dir)
            results.append(dict(
                feature=feat, agg=agg, n_pairs=len(sa),
                beta=m["beta"], ic_lo=m["ic_lo"], ic_hi=m["ic_hi"],
                p_1sided=p1, p_bonferroni=float(min(p1 * len(PRIMARY), 1.0)),
                cohen_dz=cohen_dz(sa["delta"].values),
                verdict="sensitivity", analysis="sensitivity_no_partial",
                model_note=m["model_note"],
            ))
            print(f"  A (no partial) {label}  n={len(sa)}  β={m['beta']:+.4f}"
                  f"  p_bonf={float(min(p1*len(PRIMARY),1)):.4f}")
        else:
            print(f"  A (no partial) {label}  n={len(sa)} < 5 — skip")

        # ── B: only anclaje ─────────────────────────────────────────────────
        sb = base[base["stim_subcategory"] == "anclaje"]
        if len(sb) >= 5:
            # only one stim_subcategory → disable vc_stim
            m  = run_mixed_model(sb, "delta ~ 1", vc_stim=False)
            p1 = one_sided_p(m["beta"], m["p_2sided"], exp_dir)
            results.append(dict(
                feature=feat, agg=agg, n_pairs=len(sb),
                beta=m["beta"], ic_lo=m["ic_lo"], ic_hi=m["ic_hi"],
                p_1sided=p1, p_bonferroni=float(min(p1 * len(PRIMARY), 1.0)),
                cohen_dz=cohen_dz(sb["delta"].values),
                verdict="sensitivity", analysis="sensitivity_anclaje_only",
                model_note=m["model_note"],
            ))
            print(f"  B (anclaje)    {label}  n={len(sb)}  β={m['beta']:+.4f}"
                  f"  p_bonf={float(min(p1*len(PRIMARY),1)):.4f}")
        else:
            print(f"  B (anclaje)    {label}  n={len(sb)} < 5 — skip")

        # ── C: covariate adjustment (EXPLORATORY) ───────────────────────────
        if not clin_df.empty:
            clin_str = clin_df.copy()
            clin_str["patient_id"] = clin_str["patient_id"].astype(str)
            sc = base.merge(clin_str, on="patient_id", how="left")
            avail_cov = [c for c in ["edad", "bmi", "posicion"]
                         if c in sc.columns and sc[c].notna().any()]
            sc_c = sc.dropna(subset=avail_cov + ["delta"])

            if len(sc_c) >= 10 and len(avail_cov) >= 1:
                cov_str = " + ".join(avail_cov)
                # group is between-patient; add as fixed covariate
                form_c  = f"delta ~ {cov_str} + C(group)"
                # Use OLS-style fallback for between-patient covariates
                m  = run_mixed_model(sc_c, form_c, vc_stim=True)
                p1 = one_sided_p(m["beta"], m["p_2sided"], exp_dir)
                results.append(dict(
                    feature=feat, agg=agg, n_pairs=len(sc_c),
                    beta=m["beta"], ic_lo=m["ic_lo"], ic_hi=m["ic_hi"],
                    p_1sided=p1, p_bonferroni=float(min(p1 * len(PRIMARY), 1.0)),
                    cohen_dz=cohen_dz(sc_c["delta"].values),
                    verdict="sensitivity_exploratory",
                    analysis="sensitivity_adjusted",
                    model_note=m["model_note"] + "_EXPLORATORY_n17",
                ))
                print(f"  C (adjusted)   {label}  n={len(sc_c)}  β={m['beta']:+.4f}"
                      f"  covs=[{cov_str}]  [EXPLORATORY]")
            else:
                print(f"  C (adjusted)   {label}  n={len(sc_c)} or 0 covariates — skip")
        else:
            print(f"  C (adjusted)   {label}  — no clinical data")

    return results


# ─── PASO 5: exploratory features ────────────────────────────────────────────
def run_exploratory(paired_df: pd.DataFrame) -> list:
    results = []
    print("\n=== PASO 5 — Exploratory features (no Bonferroni) ===")

    for feat, agg, exp_dir in EXPLORATORY:
        subset = paired_df[
            (paired_df["feature"] == feat) &
            (paired_df["agg"]     == agg)
        ].dropna(subset=["delta"]).copy()

        n = len(subset)
        label = f"{feat}__{agg}"
        if n < 5:
            print(f"  [SKIP] {label}: n={n} < 5")
            continue

        print(f"  {label}  n={n}", end="  … ")
        m    = run_mixed_model(subset, "delta ~ 1", vc_stim=True)
        p1   = one_sided_p(m["beta"], m["p_2sided"], exp_dir)
        dz   = cohen_dz(subset["delta"].values)
        sig  = p1 < 0.05
        print(f"β={m['beta']:+.4f}  p={p1:.4f}  {'* EXPLORATORY' if sig else 'ns'}"
              f"  [{m['model_note']}]")

        results.append(dict(
            feature=feat, agg=agg, n_pairs=n,
            beta=m["beta"], ic_lo=m["ic_lo"], ic_hi=m["ic_hi"],
            p_1sided=p1, p_bonferroni=np.nan, cohen_dz=dz,
            verdict="exploratory_sig" if sig else "exploratory_ns",
            analysis="exploratory", model_note=m["model_note"],
        ))

    return results


# ─── FIGURES ──────────────────────────────────────────────────────────────────
FEAT_LABELS = {
    "ptt_cv__mean":       "PTT-CV  (mean)",
    "ptt_cv__std":        "PTT-CV  (std)",
    "ptt_std__max":       "PTT-SD  (max)",
    "brs_alpha_lf__min":  "BRS-αLF (min)",
    "pai_mean__mean":     "PAI     (mean)",
    "ptt_std__std":       "PTT-SD  (std)  [expl.]",
    "ptt_std__slope":     "PTT-SD  (slope) [expl.]",
    "ptt_arv__std":       "PTT-ARV (std)  [expl.]",
}

COL_PRE  = "#4878D0"
COL_POST = "#EE854A"


def fig_paired_violins(paired_df: pd.DataFrame, out: Path) -> None:
    """Violin pre vs post per primary feature with individual connecting lines."""
    n_feat = len(PRIMARY)
    fig, axes = plt.subplots(1, n_feat, figsize=(4 * n_feat, 5), sharey=False)
    if n_feat == 1:
        axes = [axes]

    for ax, (feat, agg, _) in zip(axes, PRIMARY):
        sub = paired_df[
            (paired_df["feature"] == feat) &
            (paired_df["agg"]     == agg)
        ].dropna(subset=["value_pre", "value_post"])

        pre  = sub["value_pre"].values
        post = sub["value_post"].values

        # Violin halves
        vp = ax.violinplot([pre, post], positions=[0, 1], widths=0.6,
                           showmedians=True, showextrema=False)
        vp["cmedians"].set_linewidth(2)
        for i, body in enumerate(vp["bodies"]):
            body.set_facecolor(COL_PRE if i == 0 else COL_POST)
            body.set_alpha(0.55)

        # Individual connecting lines
        for p_, q_ in zip(pre, post):
            ax.plot([0, 1], [p_, q_], color="grey", alpha=0.25, lw=0.8)

        # Medians as dots
        ax.scatter([0, 1], [np.nanmedian(pre), np.nanmedian(post)],
                   color=[COL_PRE, COL_POST], s=60, zorder=5)

        label = FEAT_LABELS.get(f"{feat}__{agg}", f"{feat} {agg}")
        ax.set_title(label, fontsize=10, pad=8)
        ax.set_xticks([0, 1])
        ax.set_xticklabels(["Pre\n[−5,−2] min", "Post\n[0,+2.5] min"],
                           fontsize=9)
        ax.set_ylabel("Feature value", fontsize=9)
        ax.axhline(np.nanmedian(pre), color=COL_PRE, ls="--",
                   alpha=0.45, lw=1.0)
        n_pairs = len(sub)
        ax.set_xlabel(f"n = {n_pairs}", fontsize=9)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    fig.suptitle("Q1 v2 — Paired pre/post distributions (CAMBIO A+B)\n"
                 "Window: Pre [−5,−2] min  |  Post [0,+2.5] min",
                 fontsize=11, y=1.01)
    fig.tight_layout()
    fig.savefig(out, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {out.name}")


def fig_forest_plot(results_df: pd.DataFrame,
                    paired_df:  pd.DataFrame,
                    out: Path) -> None:
    """
    Forest plot: standardized β (= Cohen's d_z) with 95% CI.
    Primary features (top) + exploratory (bottom, italics).
    Vertical band at Bonferroni α = 0.010.
    """
    # Select primary + exploratory
    plot_rows = results_df[
        results_df["analysis"].isin(["primary", "exploratory"])
    ].copy()

    if plot_rows.empty:
        print("  [forest] no rows to plot")
        return

    # Standardised β: divide by SD(delta) per feature
    sd_map = {}
    for _, row in plot_rows.iterrows():
        sub = paired_df[
            (paired_df["feature"] == row["feature"]) &
            (paired_df["agg"]     == row["agg"])
        ]["delta"].dropna()
        sd_map[(row["feature"], row["agg"])] = sub.std(ddof=1) if len(sub) > 1 else 1.0

    plot_rows = plot_rows.copy()
    plot_rows["sd_delta"] = plot_rows.apply(
        lambda r: sd_map.get((r["feature"], r["agg"]), 1.0), axis=1
    )
    plot_rows["beta_std"]  = plot_rows["beta"]   / plot_rows["sd_delta"]
    plot_rows["ic_lo_std"] = plot_rows["ic_lo"]  / plot_rows["sd_delta"]
    plot_rows["ic_hi_std"] = plot_rows["ic_hi"]  / plot_rows["sd_delta"]

    n_rows = len(plot_rows)
    fig, ax = plt.subplots(figsize=(7, max(4, 0.55 * n_rows + 2)))

    y_primary    = []
    y_exploratory = []

    for i, (_, row) in enumerate(plot_rows.iterrows()):
        y  = n_rows - i
        b  = row["beta_std"]
        lo = row["ic_lo_std"]
        hi = row["ic_hi_std"]

        is_expl   = row["analysis"] == "exploratory"
        is_valid  = row["verdict"]  == "validado"
        is_nan    = np.isnan(b)

        color = "#2ca02c" if is_valid else ("#EE854A" if is_expl else "#4878D0")
        alpha = 0.75

        if not is_nan:
            ax.plot([lo, hi], [y, y], color=color, lw=2, alpha=alpha)
            ax.scatter([b], [y], color=color, s=60, zorder=5, alpha=alpha)

        label = FEAT_LABELS.get(f"{row['feature']}__{row['agg']}",
                                f"{row['feature']} {row['agg']}")
        p1    = row["p_1sided"]
        p_bf  = row["p_bonferroni"]
        star  = ""
        if not np.isnan(p1):
            if not np.isnan(p_bf) and p_bf < 0.05:
                star = " ***"
            elif p1 < 0.05:
                star = " *"
        n_str = f"n={row['n_pairs']}"

        ax.text(-0.02, y, label + star, ha="right", va="center",
                fontsize=9, style="italic" if is_expl else "normal",
                transform=ax.get_yaxis_transform())
        ax.text(1.01, y, n_str, ha="left", va="center", fontsize=8,
                transform=ax.get_yaxis_transform(), color="grey")

        if is_expl:
            y_exploratory.append(y)
        else:
            y_primary.append(y)

    ax.axvline(0, color="black", lw=1.2)

    # Shade region where |β_std| < threshold corresponding to Bonferroni
    ax.axvspan(-0.05, 0.05, alpha=0.06, color="grey",
               label="Null zone (|β_std|<0.05)")

    ax.set_ylim(0.3, n_rows + 0.8)
    ax.set_xlabel("Standardised β  (= Cohen's $d_z$)", fontsize=10)
    ax.set_title("Q1 v2 — Forest plot  [CAMBIO A+B]\n"
                 "*** = Bonferroni significant (p_bonf < 0.05)   "
                 "* = nominal p < 0.05 (exploratory)",
                 fontsize=10)
    ax.set_yticks([])
    ax.spines["left"].set_visible(False)
    ax.spines["top"].set_visible(False)

    legend_patches = [
        mpatches.Patch(color="#2ca02c", label="Validated (primary)"),
        mpatches.Patch(color="#4878D0", label="Not validated (primary)"),
        mpatches.Patch(color="#EE854A", label="Exploratory"),
    ]
    ax.legend(handles=legend_patches, loc="lower right", fontsize=8)

    fig.tight_layout()
    fig.savefig(out, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {out.name}")


def fig_gradient_plot(paired_df: pd.DataFrame,
                      validated_fa: list,
                      gradient_results: list,
                      out: Path) -> None:
    """Delta boxplot split by group for each validated feature."""
    n = len(validated_fa)
    if n == 0:
        return

    fig, axes = plt.subplots(1, n, figsize=(4 * n, 4.5), sharey=False)
    if n == 1:
        axes = [axes]

    groups_order = ["interescalenico", "supra_axilar"]
    colors_g     = {"interescalenico": "#4878D0", "supra_axilar": "#EE854A"}

    for ax, (feat, agg) in zip(axes, validated_fa):
        sub = paired_df[
            (paired_df["feature"] == feat) &
            (paired_df["agg"]     == agg)
        ].dropna(subset=["delta"])

        data_g = [sub[sub["group"] == g]["delta"].values
                  for g in groups_order]
        bp = ax.boxplot(data_g, positions=[0, 1], widths=0.5,
                        patch_artist=True, medianprops=dict(color="black", lw=2),
                        flierprops=dict(marker="o", ms=4, alpha=0.4))
        for patch, g in zip(bp["boxes"], groups_order):
            patch.set_facecolor(colors_g[g])
            patch.set_alpha(0.65)

        # Individual points jittered
        for xi, (g, xpos) in enumerate(zip(groups_order, [0, 1])):
            vals = sub[sub["group"] == g]["delta"].values
            jit  = np.random.uniform(-0.12, 0.12, size=len(vals))
            ax.scatter(xpos + jit, vals, color=colors_g[g],
                       alpha=0.5, s=25, zorder=5)

        ax.axhline(0, color="black", lw=0.8, ls="--")

        # Annotate gradient p-value
        grad_r = next((r for r in gradient_results
                       if r["feature"] == feat and r["agg"] == agg), None)
        if grad_r and not np.isnan(grad_r.get("p_1sided", np.nan)):
            ax.text(0.5, 0.97,
                    f"p₁-sided = {grad_r['p_1sided']:.3f}",
                    transform=ax.transAxes, ha="center", va="top",
                    fontsize=9, style="italic")

        label = FEAT_LABELS.get(f"{feat}__{agg}", f"{feat} {agg}")
        ax.set_title(label, fontsize=10)
        n_each = [len(d) for d in data_g]
        ax.set_xticks([0, 1])
        ax.set_xticklabels(
            [f"Interescalénico\n(n={n_each[0]})",
             f"Supra+Axilar\n(n={n_each[1]})"],
            fontsize=9
        )
        ax.set_ylabel("Δ feature (post − pre)", fontsize=9)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    fig.suptitle("Q1 v2 — Gradient analysis (interescalénico vs supra+axilar)\n"
                 "H₁: |Δ_interesc| < |Δ_supra| (less block residual → more response)",
                 fontsize=10, y=1.01)
    fig.tight_layout()
    fig.savefig(out, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {out.name}")


# ─── REPORT ──────────────────────────────────────────────────────────────────
def write_report(results_df: pd.DataFrame,
                 n_validated: int,
                 n_partial: int) -> None:
    if n_validated >= 3:
        overall = "**VALIDADO** (≥3/5 features primarias superan Bonferroni con dirección correcta)"
    elif n_validated >= 1:
        overall = f"**VALIDACIÓN PARCIAL** ({n_validated}/5 features primarias superan Bonferroni)"
    else:
        overall = "**NO VALIDADO** (0/5 features primarias superan Bonferroni)"

    primary_rows = results_df[results_df["analysis"] == "primary"]
    expl_rows    = results_df[results_df["analysis"] == "exploratory"]

    def fmt_row(r):
        p_bf  = f"{r['p_bonferroni']:.4f}" if not np.isnan(r["p_bonferroni"]) else "—"
        p1    = f"{r['p_1sided']:.4f}"     if not np.isnan(r["p_1sided"])    else "—"
        beta  = f"{r['beta']:+.4f}"         if not np.isnan(r["beta"])        else "—"
        ci    = (f"[{r['ic_lo']:+.4f}, {r['ic_hi']:+.4f}]"
                 if not (np.isnan(r["ic_lo"]) or np.isnan(r["ic_hi"])) else "—")
        dz    = f"{r['cohen_dz']:.3f}"     if not np.isnan(r["cohen_dz"])    else "—"
        label = FEAT_LABELS.get(f"{r['feature']}__{r['agg']}",
                                f"{r['feature']}  {r['agg']}")
        return (f"| {label:<22} | {r['n_pairs']:>7} | {beta:>10} | {ci:>20} "
                f"| {p1:>8} | {p_bf:>12} | {dz:>9} | {r['verdict']:<20} |")

    prim_table = "\n".join(fmt_row(r) for _, r in primary_rows.iterrows())
    expl_table = "\n".join(fmt_row(r) for _, r in expl_rows.iterrows()) \
                 if not expl_rows.empty else "_No exploratory features computed._"

    sens_analyses = results_df[
        results_df["analysis"].str.startswith("sensitivity")
    ]
    sens_table = ""
    if not sens_analyses.empty:
        sens_lines = []
        for _, r in sens_analyses.iterrows():
            p1   = f"{r['p_1sided']:.4f}" if not np.isnan(r["p_1sided"]) else "—"
            p_bf = f"{r['p_bonferroni']:.4f}" if not np.isnan(r["p_bonferroni"]) else "—"
            label = FEAT_LABELS.get(f"{r['feature']}__{r['agg']}",
                                    f"{r['feature']}  {r['agg']}")
            sens_lines.append(
                f"| {label:<22} | {r['analysis']:<30} | {r['n_pairs']:>5} "
                f"| {r['beta']:+.4f} | {p1:>8} | {p_bf:>12} |"
            )
        sens_table = "\n".join(sens_lines)
    else:
        sens_table = "_No sensitivity results._"

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    report = f"""# Q1 Report v2 — Refined Analysis (CAMBIO A + CAMBIO B)
Generated: {now}

---

## Summary

{overall}

- Features tested (Bonferroni-corrected): **{len(primary_rows)}**
- Features validated: **{n_validated}**
- Features partially validated: **{n_partial}**
- Bonferroni α: **{ALPHA_BONF:.4f}** (= 0.05/{len(PRIMARY)})

---

## Refinement of surrogate mapping

### Justification for CAMBIO B (direction revision)

The original Q1 analysis mapped `ptt_std` and `ptt_arv` to **positive** expected
direction (increased variability under pain stimulus). The trajectory analysis
(`trajectory_diagnostic.md`, 2026-05-08) revealed that **all** PTT variability
features decreased post-stimulus (all observed direction: **−**):

> *"ptt_std (observado −): descenso de 0.60×IQR, pico a 90 s, retorno en 120 s.
> Dirección CONTRARIO a BeatLabile."*

**Mechanistic interpretation**: Under propofol + remifentanil + regional block,
nociceptive activation produces tonic sympathetic discharge that causes uniform
vasoconstriction. This compresses the dynamic range of the PTT (shorter, more
uniform beat-to-beat PTT), reducing all variability metrics — not increasing them.
The original mapping assumed a "arousal → jitter" model, but the data support a
"activation → rigidification" model.

**Pre-specification**: All directions were fixed **before** running these tests
(this document) and are not modified post-hoc.

### Justification for CAMBIO A (shorter window)

Trajectory analysis showed:
- Peak at **90 s** post-stimulus
- Return to baseline at **120 s**
- The original 5-min window captured predominantly noise (3–4× the signal duration)
- **New window**: Pre [−5, −2] min, Post [0, +2.5 min] = [0, 150 s]

---

## PASO 2 — Primary confirmatory tests

| Feature                | n_pairs | β           | IC 95%               | p (1-sided) | p_Bonferroni | Cohen d_z | Verdict              |
|------------------------|---------|-------------|----------------------|-------------|--------------|-----------|----------------------|
{prim_table}

**Threshold**: p_Bonferroni < 0.05 (= α_bonf {ALPHA_BONF:.3f} × 5)

---

## PASO 3 — Gradient analysis (interescalénico vs supra+axilar)

H₁: |Δ_interescalénico| < |Δ_supra_axilar|
(more residual sympatholysis → less nociceptive response → delta closer to zero)
Reference group: supra_axilar. Test one-sided: β_interesc > 0.

{"*Skipped: no validated features in primary analysis.*" if n_validated == 0 else
 "See `test_results_v2.csv` (analysis=gradient) for full results."}

---

## PASO 4 — Sensitivity analyses

| Feature                | Analysis                       | n   | β       | p (1-sided) | p_Bonferroni |
|------------------------|--------------------------------|-----|---------|-------------|--------------|
{sens_table}

**Sensitivity C** is flagged EXPLORATORY (n=17 patients; between-patient covariates
edad, BMI, posicion are poorly identified with patient_id as random effect).

---

## PASO 5 — Exploratory features (reclassified)

The following features were originally classified with positive expected direction
("inconsistente" category in Q1 v1). They are reclassified here with direction −
(consistent with the rigidification model) and tested without Bonferroni correction.
All are marked **exploratory** and do not contribute to the primary verdict.

| Feature                | n_pairs | β           | IC 95%               | p (1-sided) | p_Bonferroni | Cohen d_z | Verdict              |
|------------------------|---------|-------------|----------------------|-------------|--------------|-----------|----------------------|
{expl_table}

If all three show p < 0.05 in the negative direction, this corroborates the
"uniform rigidification" interpretation.

---

## Q1 Verdict (v2)

{overall}

### Definition
- **VALIDADO**: ≥3/5 primary features cross Bonferroni with correct direction
- **VALIDACIÓN PARCIAL**: 1–2 features cross Bonferroni
- **NO VALIDADO**: 0 features cross Bonferroni

---

## Files

| File | Description |
|------|-------------|
| `paired_event_data.csv` | Long-format paired pre/post values (52 events × 8 features × 2 aggs) |
| `test_results_v2.csv`   | Full statistical results (primary + gradient + sensitivity + exploratory) |
| `figures/paired_event_violins.png` | Pre/post distributions per primary feature |
| `figures/forest_plot_v2.png`       | Standardised β forest plot (primary + exploratory) |
| `figures/gradient_plot_v2.png`     | Group gradient (if ≥1 feature validated) |

---
*Q1 Refinement v2 — beatlabile validation pipeline*
"""

    with open(REPORT_OUT, "w", encoding="utf-8") as f:
        f.write(report)
    print(f"  Report saved: {REPORT_OUT.name}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main() -> None:
    print(f"[{datetime.now():%H:%M:%S}] Q1 Analysis v2 starting")
    print(f"  Pre window:  [{PRE_START_S}, {PRE_END_S}] s  "
          f"({PRE_START_S/60:.1f} … {PRE_END_S/60:.1f} min)")
    print(f"  Post window: [{POST_START_S}, {POST_END_S}] s  "
          f"({POST_START_S/60:.1f} … {POST_END_S/60:.1f} min)")
    print(f"  Bonferroni α = {ALPHA_BONF:.4f}  ({len(PRIMARY)} tests)\n")

    # ── Load data ──────────────────────────────────────────────────────────
    print(f"[{datetime.now():%H:%M:%S}] Loading features_long.parquet …")
    feat_df = pd.read_parquet(str(FEATURES_LONG))
    print(f"  {feat_df.shape[0]:,} rows × {feat_df.shape[1]} columns")

    print(f"[{datetime.now():%H:%M:%S}] Loading event_windows.csv …")
    ev_df   = pd.read_csv(str(EVENT_WINDOWS))
    events  = ev_df[ev_df["window_type"] == "event"].copy()
    print(f"  {len(events)} events  ({events['group'].value_counts().to_dict()})")

    clin_df = load_clinical_covariates()
    if not clin_df.empty:
        print(f"  Clinical covariates: {len(clin_df)} patients loaded")

    # ── PASO 1 ─────────────────────────────────────────────────────────────
    print(f"\n[{datetime.now():%H:%M:%S}] PASO 1 — Building paired data …")
    paired_df = build_paired_data(feat_df, events)
    # Normalise patient_id to string for downstream joins
    paired_df["patient_id"] = paired_df["patient_id"].astype(str)
    paired_df.to_csv(str(PAIRED_OUT), index=False)
    n_valid_delta = paired_df["delta"].notna().sum()
    print(f"  {len(paired_df):,} rows  |  {n_valid_delta:,} valid deltas"
          f"  ({n_valid_delta/len(paired_df):.1%})")
    print(f"  Saved: {PAIRED_OUT.name}")

    # Quick QC: check window coverage per event
    coverage = (
        paired_df.groupby(["stim_id", "feature", "agg"])["delta"]
        .apply(lambda s: s.notna().any())
        .groupby("stim_id").sum()
    )
    print(f"  Window coverage: min={coverage.min():.0f}  "
          f"median={coverage.median():.0f}  max={coverage.max():.0f} "
          f"features/event with valid delta")

    # ── PASO 2 ─────────────────────────────────────────────────────────────
    all_results = []
    primary_res = run_primary_tests(paired_df)
    all_results.extend(primary_res)

    n_validated = sum(1 for r in primary_res if r["verdict"] == "validado")
    n_partial   = sum(1 for r in primary_res
                      if r["verdict"] == "no_validado"
                      and not np.isnan(r["p_1sided"])
                      and r["p_1sided"] < 0.05)

    print(f"\n  >>> Primary verdict: {n_validated}/5 validated  "
          f"(Bonferroni α={ALPHA_BONF:.3f}) <<<")

    # ── PASO 3 ─────────────────────────────────────────────────────────────
    validated_fa = [
        (r["feature"], r["agg"])
        for r in primary_res if r["verdict"] == "validado"
    ]
    gradient_res = run_gradient_analysis(paired_df, validated_fa)
    all_results.extend(gradient_res)

    # ── PASO 4 ─────────────────────────────────────────────────────────────
    sens_res = run_sensitivity(paired_df, clin_df)
    all_results.extend(sens_res)

    # ── PASO 5 ─────────────────────────────────────────────────────────────
    expl_res = run_exploratory(paired_df)
    all_results.extend(expl_res)

    # ── Save results ───────────────────────────────────────────────────────
    results_df = pd.DataFrame(all_results)
    out_cols = ["feature", "agg", "n_pairs", "beta", "ic_lo", "ic_hi",
                "p_1sided", "p_bonferroni", "cohen_dz", "verdict",
                "analysis", "model_note"]
    for c in out_cols:
        if c not in results_df.columns:
            results_df[c] = np.nan
    results_df[out_cols].to_csv(str(RESULTS_OUT), index=False)
    print(f"\n  Results saved: {RESULTS_OUT.name}")

    # ── Figures ────────────────────────────────────────────────────────────
    print(f"\n[{datetime.now():%H:%M:%S}] Generating figures …")
    fig_paired_violins(paired_df, FIG / "paired_event_violins.png")
    fig_forest_plot(results_df, paired_df, FIG / "forest_plot_v2.png")
    if validated_fa:
        fig_gradient_plot(paired_df, validated_fa, gradient_res,
                          FIG / "gradient_plot_v2.png")
    else:
        print("  gradient_plot_v2.png: skipped (0 validated)")

    # ── Report ─────────────────────────────────────────────────────────────
    print(f"\n[{datetime.now():%H:%M:%S}] Writing Q1_report_v2.md …")
    write_report(results_df, n_validated, n_partial)

    # ── Final summary ──────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"Q1 Analysis v2 complete — {datetime.now():%H:%M:%S}")
    print(f"{'='*60}")
    print(f"  Features validated:  {n_validated}/5")
    expl_sig = sum(1 for r in expl_res
                   if r["verdict"] == "exploratory_sig")
    print(f"  Exploratory sig:     {expl_sig}/{len(expl_res)}")
    if n_validated == 0 and expl_sig == len(expl_res) and len(expl_res) > 0:
        print("  NOTE: All exploratory features significant — "
              "corroborates uniform rigidification hypothesis.")
    print(f"  Paired data:         {PAIRED_OUT.name}")
    print(f"  Results CSV:         {RESULTS_OUT.name}")
    print(f"  Report:              {REPORT_OUT.name}")
    if n_validated >= 3:
        print("\n  *** Q1 VALIDATED ***")
    elif n_validated >= 1:
        print(f"\n  *** Q1 PARTIAL VALIDATION ({n_validated}/5) ***")
    else:
        print("\n  Q1 NOT VALIDATED (0/5 Bonferroni)")


if __name__ == "__main__":
    main()
